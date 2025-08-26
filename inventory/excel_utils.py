import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, date
from django.contrib.auth.models import User
from .models import MaterialType, Supplier, Customer, StockIn, StockOut, Inventory, StockTransfer
from .utils import gregorian_to_persian_str, gregorian_to_persian_datetime_str
import os

def create_unified_stock_template():
    """ایجاد قالب Excel یکپارچه برای ورودی و خروجی انبار"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ورودی و خروجی انبار"
    
    # تعریف ستون‌ها
    headers = [
        "انبار",
        "نوع عملیات",
        "نام کالا",
        "هویت کالا/نام مشتری", 
        "مقدار",
        "قیمت واحد",
        "شماره بارنامه",
        "تاریخ (YYYY-MM-DD)",
        "یادداشت‌ها"
    ]
    
    # اضافه کردن هدرها
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # تنظیم عرض ستون‌ها
    column_widths = [15, 15, 20, 25, 15, 15, 20, 20, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # اضافه کردن نمونه داده
    sample_data = [
        ["انبار اصلی", "ورودی", "میلگرد 16", "شرکت آهن آلات تهران", 1000, 15000, "BR001", "2024-01-15", "ورودی اولیه"],
        ["انبار اصلی", "خروجی", "میلگرد 16", "شرکت ساختمانی آسمان", 200, 18000, "BR002", "2024-01-16", "خروجی اولیه"],
        ["انبار اصلی", "ورودی", "ورق فولادی", "کارخانه فولاد اصفهان", 500, 25000, "BR003", "2024-01-17", "ورودی دوم"],
        ["انبار اصلی", "خروجی", "ورق فولادی", "پروژه برج تهران", 100, 28000, "BR004", "2024-01-18", "خروجی دوم"],
    ]
    
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # رنگ‌بندی بر اساس نوع عملیات
            if col == 1:  # ستون نوع عملیات
                if value == "ورودی":
                    cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                elif value == "خروجی":
                    cell.fill = PatternFill(start_color="F8E8E8", end_color="F8E8E8", fill_type="solid")
    
    # ذخیره فایل
    filename = f"قالب_یکپارچه_ورودی_خروجی_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("media", "excel_templates", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath

def import_unified_stock_excel(file_path, user):
    """وارد کردن داده‌های یکپارچه ورودی و خروجی انبار از فایل Excel"""
    try:
        df = pd.read_excel(file_path)
        results = {"success": [], "errors": []}
        
        # بررسی ستون‌های موجود و تطبیق با نام‌های مختلف
        column_mapping = {
            'انبار': ['انبار', 'warehouse'],
            'نوع عملیات': ['نوع عملیات', 'نوع عمليات', 'operation_type'],
            'نام کالا': ['نام کالا', 'نام كالا', 'material_name'],
            'هویت کالا/نام مشتری': ['هویت کالا/نام مشتری', 'هویت كالا/نام مشتری', 'هویت کالا', 'هویت كالا', 'supplier_customer'],
            'مقدار': ['مقدار', 'quantity'],
            'قیمت واحد': ['قیمت واحد', 'قيمت واحد', 'unit_price'],
            'شماره بارنامه': ['شماره بارنامه', 'شماره بارنامه', 'invoice_number'],
            'تاریخ (YYYY-MM-DD)': ['تاریخ (YYYY-MM-DD)', 'تاريخ (YYYY-MM-DD)', 'date'],
            'یادداشت‌ها': ['یادداشت‌ها', 'يادداشت‌ها', 'notes']
        }
        
        # پیدا کردن ستون‌های موجود
        found_columns = {}
        missing_columns = []
        
        for required_col, possible_names in column_mapping.items():
            found = False
            for possible_name in possible_names:
                if possible_name in df.columns:
                    found_columns[required_col] = possible_name
                    found = True
                    break
            if not found:
                missing_columns.append(required_col)
        
        if missing_columns:
            results["errors"].append(f"ستون‌های زیر در فایل یافت نشد: {', '.join(missing_columns)}")
            results["errors"].append(f"ستون‌های موجود: {', '.join(df.columns)}")
            return results
        
        for index, row in df.iterrows():
            try:
                # دریافت نوع عملیات
                operation_type = str(row[found_columns['نوع عملیات']]).strip()
                if operation_type not in ['ورودی', 'خروجی']:
                    results["errors"].append(f"ردیف {index + 2}: نوع عملیات نامعتبر - باید 'ورودی' یا 'خروجی' باشد")
                    continue
                
                # دریافت یا ایجاد نام کالا
                material_name = str(row[found_columns['نام کالا']]).strip()
                if not material_name:
                    results["errors"].append(f"ردیف {index + 2}: نام کالا خالی است")
                    continue
                    
                material_type, created = MaterialType.objects.get_or_create(
                    name=material_name,
                    defaults={'unit': 'کیلوگرم'}
                )
                
                # دریافت هویت کالا/نام مشتری
                supplier_customer_name = str(row[found_columns['هویت کالا/نام مشتری']]).strip()
                if not supplier_customer_name:
                    results["errors"].append(f"ردیف {index + 2}: هویت کالا/نام مشتری خالی است")
                    continue
                
                # تبدیل داده‌ها
                quantity = int(row[found_columns['مقدار']]) if pd.notna(row[found_columns['مقدار']]) else 0
                unit_price = int(row[found_columns['قیمت واحد']]) if pd.notna(row[found_columns['قیمت واحد']]) else 0
                invoice_number = str(row[found_columns['شماره بارنامه']]).strip() if pd.notna(row[found_columns['شماره بارنامه']]) else ""
                notes = str(row[found_columns['یادداشت‌ها']]).strip() if pd.notna(row[found_columns['یادداشت‌ها']]) else ""
                
                # تبدیل تاریخ
                manual_date = None
                if pd.notna(row[found_columns['تاریخ (YYYY-MM-DD)']]):
                    try:
                        if isinstance(row[found_columns['تاریخ (YYYY-MM-DD)']], str):
                            manual_date = datetime.strptime(row[found_columns['تاریخ (YYYY-MM-DD)']], '%Y-%m-%d').date()
                        else:
                            manual_date = row[found_columns['تاریخ (YYYY-MM-DD)']].date()
                    except:
                        manual_date = None
                
                # دریافت انبار
                warehouse_name = str(row[found_columns['انبار']]).strip() if 'انبار' in found_columns and pd.notna(row[found_columns['انبار']]) else "انبار اصلی"
                warehouse, created = Warehouse.objects.get_or_create(
                    name=warehouse_name,
                    defaults={'code': warehouse_name[:10].upper(), 'is_active': True}
                )
                
                if operation_type == "ورودی":
                    # پردازش ورودی انبار
                    supplier, created = Supplier.objects.get_or_create(
                        name=supplier_customer_name
                    )
                    
                    # ایجاد رکورد ورودی
                    stock_in = StockIn.objects.create(
                        warehouse=warehouse,
                        material_type=material_type,
                        supplier=supplier,
                        quantity=quantity,
                        unit_price=unit_price,
                        invoice_number=invoice_number,
                        notes=notes,
                        created_by=user,
                        manual_date=manual_date
                    )
                    
                    # بروزرسانی موجودی
                    inventory, created = Inventory.objects.get_or_create(
                        warehouse=warehouse,
                        material_type=material_type,
                        defaults={'current_quantity': 0}
                    )
                    inventory.current_quantity += quantity
                    inventory.save()
                    
                    results["success"].append(f"ردیف {index + 2}: ورودی {material_name} با موفقیت ثبت شد")
                    
                elif operation_type == "خروجی":
                    # پردازش خروجی انبار
                    customer, created = Customer.objects.get_or_create(
                        name=supplier_customer_name
                    )
                    
                    # بررسی موجودی
                    try:
                        inventory = Inventory.objects.get(warehouse=warehouse, material_type=material_type)
                        if inventory.current_quantity < quantity:
                            results["errors"].append(f"ردیف {index + 2}: موجودی ناکافی برای {material_name} در انبار {warehouse.name} (موجودی: {inventory.current_quantity}, درخواستی: {quantity})")
                            continue
                    except Inventory.DoesNotExist:
                        results["errors"].append(f"ردیف {index + 2}: موجودی برای {material_name} در انبار {warehouse.name} یافت نشد")
                        continue
                    
                    # ایجاد رکورد خروجی
                    stock_out = StockOut.objects.create(
                        warehouse=warehouse,
                        material_type=material_type,
                        customer=customer,
                        quantity=quantity,
                        unit_price=unit_price,
                        invoice_number=invoice_number,
                        notes=notes,
                        created_by=user,
                        manual_date=manual_date
                    )
                    
                    # بروزرسانی موجودی
                    inventory.current_quantity -= quantity
                    inventory.save()
                    
                    results["success"].append(f"ردیف {index + 2}: خروجی {material_name} با موفقیت ثبت شد")
                
            except Exception as e:
                results["errors"].append(f"ردیف {index + 2}: خطا - {str(e)}")
        
        return results
        
    except Exception as e:
        return {"success": [], "errors": [f"خطا در خواندن فایل: {str(e)}"]}

def create_stock_in_template():
    """ایجاد قالب Excel برای ورودی انبار"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ورودی انبار"
    
    # تعریف ستون‌ها
    headers = [
        "انبار",
        "نام کالا",
        "هویت کالا", 
        "مقدار",
        "قیمت واحد",
        "شماره بارنامه",
        "تاریخ ورود (YYYY-MM-DD)",
        "یادداشت‌ها"
    ]
    
    # اضافه کردن هدرها
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # تنظیم عرض ستون‌ها
    column_widths = [15, 20, 25, 15, 15, 20, 20, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # اضافه کردن نمونه داده
    sample_data = [
        ["انبار اصلی", "میلگرد 16", "شرکت آهن آلات تهران", 1000, 15000, "BR001", "2024-01-15", "ورودی اولیه"],
        ["انبار اصلی", "ورق فولادی", "کارخانه فولاد اصفهان", 500, 25000, "BR002", "2024-01-16", "ورودی دوم"],
    ]
    
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # ذخیره فایل
    filename = f"قالب_ورودی_انبار_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("media", "excel_templates", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath

def create_stock_out_template():
    """ایجاد قالب Excel برای خروجی انبار"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "خروجی انبار"
    
    # تعریف ستون‌ها
    headers = [
        "انبار",
        "نام کالا",
        "نام مشتری", 
        "مقدار",
        "قیمت واحد",
        "شماره بارنامه",
        "تاریخ خروج (YYYY-MM-DD)",
        "یادداشت‌ها"
    ]
    
    # اضافه کردن هدرها
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # تنظیم عرض ستون‌ها
    column_widths = [15, 20, 25, 15, 15, 20, 20, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # اضافه کردن نمونه داده
    sample_data = [
        ["انبار اصلی", "میلگرد 16", "شرکت ساختمانی آسمان", 200, 18000, "BR003", "2024-01-17", "خروجی اولیه"],
        ["انبار اصلی", "ورق فولادی", "پروژه برج تهران", 100, 28000, "BR004", "2024-01-18", "خروجی دوم"],
    ]
    
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # ذخیره فایل
    filename = f"قالب_خروجی_انبار_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("media", "excel_templates", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath

def import_stock_in_excel(file_path, user):
    """وارد کردن داده‌های ورودی انبار از فایل Excel"""
    try:
        df = pd.read_excel(file_path)
        results = {"success": [], "errors": []}
        
        for index, row in df.iterrows():
            try:
                # دریافت انبار
                warehouse_name = str(row['انبار']).strip() if 'انبار' in df.columns and pd.notna(row['انبار']) else "انبار اصلی"
                warehouse, created = Warehouse.objects.get_or_create(
                    name=warehouse_name,
                    defaults={'code': warehouse_name[:10].upper(), 'is_active': True}
                )
                
                # دریافت یا ایجاد نام کالا
                material_name = str(row['نام کالا']).strip()
                material_type, created = MaterialType.objects.get_or_create(
                    name=material_name,
                    defaults={'unit': 'کیلوگرم'}
                )
                
                # دریافت یا ایجاد هویت کالا
                supplier_name = str(row['هویت کالا']).strip()
                supplier, created = Supplier.objects.get_or_create(
                    name=supplier_name
                )
                
                # تبدیل داده‌ها
                quantity = int(row['مقدار']) if pd.notna(row['مقدار']) else 0
                unit_price = int(row['قیمت واحد']) if pd.notna(row['قیمت واحد']) else 0
                invoice_number = str(row['شماره بارنامه']).strip() if pd.notna(row['شماره بارنامه']) else ""
                notes = str(row['یادداشت‌ها']).strip() if pd.notna(row['یادداشت‌ها']) else ""
                
                # تبدیل تاریخ
                manual_date = None
                if pd.notna(row['تاریخ ورود (YYYY-MM-DD)']):
                    try:
                        if isinstance(row['تاریخ ورود (YYYY-MM-DD)'], str):
                            manual_date = datetime.strptime(row['تاریخ ورود (YYYY-MM-DD)'], '%Y-%m-%d').date()
                        else:
                            manual_date = row['تاریخ ورود (YYYY-MM-DD)'].date()
                    except:
                        manual_date = None
                
                # ایجاد رکورد ورودی
                stock_in = StockIn.objects.create(
                    warehouse=warehouse,
                    material_type=material_type,
                    supplier=supplier,
                    quantity=quantity,
                    unit_price=unit_price,
                    invoice_number=invoice_number,
                    notes=notes,
                    created_by=user,
                    manual_date=manual_date
                )
                
                results["success"].append(f"ردیف {index + 2}: ورودی {material_name} با موفقیت ثبت شد")
                
            except Exception as e:
                results["errors"].append(f"ردیف {index + 2}: خطا - {str(e)}")
        
        return results
        
    except Exception as e:
        return {"success": [], "errors": [f"خطا در خواندن فایل: {str(e)}"]}

def import_stock_out_excel(file_path, user):
    """وارد کردن داده‌های خروجی انبار از فایل Excel"""
    try:
        df = pd.read_excel(file_path)
        results = {"success": [], "errors": []}
        
        for index, row in df.iterrows():
            try:
                # دریافت انبار
                warehouse_name = str(row['انبار']).strip() if 'انبار' in df.columns and pd.notna(row['انبار']) else "انبار اصلی"
                warehouse, created = Warehouse.objects.get_or_create(
                    name=warehouse_name,
                    defaults={'code': warehouse_name[:10].upper(), 'is_active': True}
                )
                
                # دریافت یا ایجاد نام کالا
                material_name = str(row['نام کالا']).strip()
                material_type, created = MaterialType.objects.get_or_create(
                    name=material_name,
                    defaults={'unit': 'کیلوگرم'}
                )
                
                # دریافت یا ایجاد مشتری
                customer_name = str(row['نام مشتری']).strip()
                customer, created = Customer.objects.get_or_create(
                    name=customer_name
                )
                
                # تبدیل داده‌ها
                quantity = int(row['مقدار']) if pd.notna(row['مقدار']) else 0
                unit_price = int(row['قیمت واحد']) if pd.notna(row['قیمت واحد']) else 0
                invoice_number = str(row['شماره بارنامه']).strip() if pd.notna(row['شماره بارنامه']) else ""
                notes = str(row['یادداشت‌ها']).strip() if pd.notna(row['یادداشت‌ها']) else ""
                
                # تبدیل تاریخ
                manual_date = None
                if pd.notna(row['تاریخ خروج (YYYY-MM-DD)']):
                    try:
                        if isinstance(row['تاریخ خروج (YYYY-MM-DD)'], str):
                            manual_date = datetime.strptime(row['تاریخ خروج (YYYY-MM-DD)'], '%Y-%m-%d').date()
                        else:
                            manual_date = row['تاریخ خروج (YYYY-MM-DD)'].date()
                    except:
                        manual_date = None
                
                # بررسی موجودی
                try:
                    inventory = Inventory.objects.get(warehouse=warehouse, material_type=material_type)
                    if inventory.current_quantity < quantity:
                        results["errors"].append(f"ردیف {index + 2}: موجودی ناکافی برای {material_name} در انبار {warehouse.name}")
                        continue
                except Inventory.DoesNotExist:
                    results["errors"].append(f"ردیف {index + 2}: موجودی برای {material_name} در انبار {warehouse.name} یافت نشد")
                    continue
                
                # ایجاد رکورد خروجی
                stock_out = StockOut.objects.create(
                    warehouse=warehouse,
                    material_type=material_type,
                    customer=customer,
                    quantity=quantity,
                    unit_price=unit_price,
                    invoice_number=invoice_number,
                    notes=notes,
                    created_by=user,
                    manual_date=manual_date
                )
                
                results["success"].append(f"ردیف {index + 2}: خروجی {material_name} با موفقیت ثبت شد")
                
            except Exception as e:
                results["errors"].append(f"ردیف {index + 2}: خطا - {str(e)}")
        
        return results
        
    except Exception as e:
        return {"success": [], "errors": [f"خطا در خواندن فایل: {str(e)}"]}

def create_stock_transfer_template():
    """ایجاد قالب Excel برای انتقال انبار"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "انتقال انبار"
    
    # تعریف ستون‌ها
    headers = [
        "نام کالا",
        "نوع انتقال", 
        "مقدار",
        "مکان مبدا",
        "مکان مقصد",
        "تاریخ انتقال (YYYY-MM-DD)",
        "یادداشت‌ها"
    ]
    
    # اضافه کردن هدرها
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # تنظیم عرض ستون‌ها
    column_widths = [20, 15, 15, 20, 20, 20, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # اضافه کردن نمونه داده
    sample_data = [
        ["میلگرد 16", "انتقال به انبار", 200, "انبار اصلی", "انبار فرعی", "2024-01-20", "انتقال اولیه"],
        ["ورق فولادی", "انتقال از انبار", 100, "انبار فرعی", "انبار اصلی", "2024-01-21", "انتقال دوم"],
    ]
    
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # ذخیره فایل
    filename = f"قالب_انتقال_انبار_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("media", "excel_templates", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath

def import_stock_transfer_excel(file_path, user):
    """وارد کردن داده‌های انتقال انبار از فایل Excel"""
    try:
        df = pd.read_excel(file_path)
        results = {"success": [], "errors": []}
        
        # بررسی ستون‌های موجود
        required_columns = [
            'نام کالا',
            'نوع انتقال',
            'مقدار',
            'مکان مبدا',
            'مکان مقصد',
            'تاریخ انتقال (YYYY-MM-DD)',
            'یادداشت‌ها'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            results["errors"].append(f"ستون‌های زیر در فایل یافت نشد: {', '.join(missing_columns)}")
            results["errors"].append(f"ستون‌های موجود: {', '.join(df.columns)}")
            return results
        
        for index, row in df.iterrows():
            try:
                # دریافت یا ایجاد نام کالا
                material_name = str(row['نام کالا']).strip()
                if not material_name:
                    results["errors"].append(f"ردیف {index + 2}: نام کالا خالی است")
                    continue
                    
                material_type, created = MaterialType.objects.get_or_create(
                    name=material_name,
                    defaults={'unit': 'کیلوگرم'}
                )
                
                # دریافت نوع انتقال
                transfer_type_str = str(row['نوع انتقال']).strip()
                if 'به انبار' in transfer_type_str:
                    transfer_type = 'in'
                elif 'از انبار' in transfer_type_str:
                    transfer_type = 'out'
                else:
                    results["errors"].append(f"ردیف {index + 2}: نوع انتقال نامعتبر - باید شامل 'به انبار' یا 'از انبار' باشد")
                    continue
                
                # دریافت سایر داده‌ها
                quantity = int(row['مقدار']) if pd.notna(row['مقدار']) else 0
                source_location = str(row['مکان مبدا']).strip() if pd.notna(row['مکان مبدا']) else ""
                destination_location = str(row['مکان مقصد']).strip() if pd.notna(row['مکان مقصد']) else ""
                notes = str(row['یادداشت‌ها']).strip() if pd.notna(row['یادداشت‌ها']) else ""
                
                # تبدیل تاریخ
                transfer_date = None
                if pd.notna(row['تاریخ انتقال (YYYY-MM-DD)']):
                    try:
                        if isinstance(row['تاریخ انتقال (YYYY-MM-DD)'], str):
                            transfer_date = datetime.strptime(row['تاریخ انتقال (YYYY-MM-DD)'], '%Y-%m-%d').date()
                        else:
                            transfer_date = row['تاریخ انتقال (YYYY-MM-DD)'].date()
                    except:
                        transfer_date = None
                
                # ایجاد رکورد انتقال
                stock_transfer = StockTransfer.objects.create(
                    material_type=material_type,
                    transfer_type=transfer_type,
                    quantity=quantity,
                    source_location=source_location,
                    destination_location=destination_location,
                    notes=notes,
                    created_by=user
                )
                
                results["success"].append(f"ردیف {index + 2}: انتقال {material_name} با موفقیت ثبت شد")
                
            except Exception as e:
                results["errors"].append(f"ردیف {index + 2}: خطا - {str(e)}")
        
        return results
        
    except Exception as e:
        return {"success": [], "errors": [f"خطا در خواندن فایل: {str(e)}"]}

def export_inventory_to_excel():
    """صدور موجودی انبار به فایل Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "موجودی انبار"
    
    # تعریف هدرها
    headers = [
        "نام کالا",
        "واحد اندازه‌گیری",
        "موجودی فعلی",
        "آخرین بروزرسانی"
    ]
    
    # اضافه کردن هدرها
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # تنظیم عرض ستون‌ها
    column_widths = [25, 20, 15, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # اضافه کردن داده‌ها
    inventories = Inventory.objects.all().select_related('material_type')
    for row, inventory in enumerate(inventories, 2):
        ws.cell(row=row, column=1, value=inventory.material_type.name)
        ws.cell(row=row, column=2, value=inventory.material_type.unit)
        ws.cell(row=row, column=3, value=inventory.current_quantity or 0)
        ws.cell(row=row, column=4, value=gregorian_to_persian_datetime_str(inventory.last_updated, "%Y/%m/%d %H:%M"))
    
    # ذخیره فایل
    filename = f"موجودی_انبار_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("media", "excel_reports", filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)
    return filepath
