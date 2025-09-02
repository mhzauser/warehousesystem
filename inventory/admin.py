from django.contrib import admin
from django.http import HttpResponse
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils.html import format_html
import os
import tempfile
from datetime import datetime
from .models import Warehouse, MaterialType, Supplier, Customer, Inventory, StockIn, StockOut, StockTransfer
from .excel_utils import (
    create_stock_in_template, create_stock_out_template, 
    import_stock_in_excel, import_stock_out_excel,
    export_inventory_to_excel, create_stock_transfer_template,
    import_stock_transfer_excel
)
from .utils import gregorian_to_persian_str, gregorian_to_persian_datetime_str

@admin.register(Warehouse)
class WarehouseAdmin(admin.ModelAdmin):
    list_display = ['name', 'code', 'manager', 'phone', 'is_active', 'persian_created_at']
    list_filter = ['is_active', 'created_at']
    search_fields = ['name', 'code', 'manager', 'phone']
    readonly_fields = ['created_at']
    ordering = ['name']
    
    def persian_created_at(self, obj):
        return gregorian_to_persian_str(obj.created_at, "%Y/%m/%d")
    persian_created_at.short_description = 'تاریخ ایجاد (شمسی)'

@admin.register(MaterialType)
class MaterialTypeAdmin(admin.ModelAdmin):
    list_display = ['name', 'unit', 'description']
    search_fields = ['name']
    list_filter = ['unit']

@admin.register(Supplier)
class SupplierAdmin(admin.ModelAdmin):
    list_display = ['name', 'contact_person', 'phone', 'address', 'persian_created_at']
    search_fields = ['name', 'contact_person', 'phone']
    list_filter = ['created_at']
    
    def persian_created_at(self, obj):
        return gregorian_to_persian_str(obj.created_at, "%Y/%m/%d")
    persian_created_at.short_description = 'تاریخ ایجاد (شمسی)'

@admin.register(Customer)
class CustomerAdmin(admin.ModelAdmin):
    list_display = ['name', 'contact_person', 'phone', 'address', 'persian_created_at']
    search_fields = ['name', 'contact_person', 'phone']
    list_filter = ['created_at']
    
    def persian_created_at(self, obj):
        return gregorian_to_persian_str(obj.created_at, "%Y/%m/%d")
    persian_created_at.short_description = 'تاریخ ایجاد (شمسی)'

@admin.register(Inventory)
class InventoryAdmin(admin.ModelAdmin):
    list_display = ['warehouse', 'material_type', 'supplier', 'current_quantity', 'unit_display', 'persian_last_updated']
    list_filter = ['warehouse', 'material_type', 'supplier', 'last_updated']
    search_fields = ['warehouse__name', 'material_type__name', 'supplier__name']
    readonly_fields = ['last_updated']
    actions = ['export_inventory_excel']
    
    def persian_last_updated(self, obj):
        return gregorian_to_persian_datetime_str(obj.last_updated, "%Y/%m/%d %H:%M")
    persian_last_updated.short_description = 'آخرین بروزرسانی (شمسی)'
    
    def unit_display(self, obj):
        return obj.material_type.unit
    unit_display.short_description = 'واحد'
    
    def export_inventory_excel(self, request, queryset):
        try:
            file_path = export_inventory_to_excel()
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
                return response
        except Exception as e:
            messages.error(request, f'خطا در ایجاد گزارش: {str(e)}')
            return redirect('admin:inventory_inventory_changelist')
    
    export_inventory_excel.short_description = "صدور موجودی به Excel"

@admin.register(StockIn)
class StockInAdmin(admin.ModelAdmin):
    list_display = ['warehouse', 'material_type', 'supplier', 'customer', 'quantity', 'unit_price', 'total_price', 'persian_manual_date', 'persian_created_at']
    list_filter = ['warehouse', 'material_type', 'supplier', 'customer', 'created_at', 'manual_date']
    search_fields = ['warehouse__name', 'material_type__name', 'supplier__name', 'customer__name', 'invoice_number']
    readonly_fields = ['total_price', 'created_at']
    date_hierarchy = 'created_at'
    actions = ['export_stock_in_excel']
    
    def persian_manual_date(self, obj):
        if obj.manual_date:
            return gregorian_to_persian_str(obj.manual_date, "%Y/%m/%d")
        return "-"
    persian_manual_date.short_description = 'تاریخ ورود دستی (شمسی)'
    
    def persian_created_at(self, obj):
        return gregorian_to_persian_datetime_str(obj.created_at, "%Y/%m/%d %H:%M")
    persian_created_at.short_description = 'تاریخ ثبت (شمسی)'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel_view), name='inventory_stockin_upload_excel'),
            path('download-template/', self.admin_site.admin_view(self.download_template_view), name='inventory_stockin_download_template'),
        ]
        return custom_urls + urls
    
    def upload_excel_view(self, request):
        if request.method == 'POST':
            uploaded_file = request.FILES.get('excel_file')
            if uploaded_file:
                try:
                    # ذخیره فایل موقت
                    temp_path = f'/tmp/stock_in_admin_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                    with open(temp_path, 'wb+') as destination:
                        for chunk in uploaded_file.chunks():
                            destination.write(chunk)
                    
                    # پردازش فایل
                    results = import_stock_in_excel(temp_path, request.user)
                    
                    # حذف فایل موقت
                    os.remove(temp_path)
                    
                    # نمایش نتایج
                    success_count = len(results.get('success', []))
                    error_count = len(results.get('errors', []))
                    
                    if success_count > 0:
                        messages.success(request, f'{success_count} رکورد با موفقیت وارد شد.')
                    
                    if error_count > 0:
                        for error in results['errors'][:5]:  # نمایش 5 خطای اول
                            messages.error(request, error)
                        if error_count > 5:
                            messages.warning(request, f'و {error_count - 5} خطای دیگر وجود دارد.')
                    
                    return redirect('admin:inventory_stockin_changelist')
                    
                except Exception as e:
                    messages.error(request, f'خطا در پردازش فایل: {str(e)}')
            else:
                messages.error(request, 'فایل انتخاب نشده است.')
        
        return render(request, 'admin/inventory/stockin/upload_excel.html', {
            'title': 'آپلود Excel ورودی انبار',
            'opts': self.model._meta,
        })
    
    def download_template_view(self, request):
        try:
            file_path = create_stock_in_template()
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
                return response
        except Exception as e:
            messages.error(request, f'خطا در ایجاد قالب: {str(e)}')
            return redirect('admin:inventory_stockin_changelist')
    
    def export_stock_in_excel(self, request, queryset):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ورودی انبار"
            
            # هدرها
            headers = [
                "انبار", "نام کالا", "هویت کالا", "مشتری", "مقدار", "قیمت واحد", 
                "قیمت کل", "شماره بارنامه", "تاریخ ورود دستی", "یادداشت‌ها"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # داده‌ها
            for row, stock_in in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=stock_in.warehouse.name)
                ws.cell(row=row, column=2, value=stock_in.material_type.name)
                ws.cell(row=row, column=3, value=stock_in.supplier.name)
                ws.cell(row=row, column=4, value=stock_in.quantity or 0)
                ws.cell(row=row, column=5, value=stock_in.unit_price or 0)
                ws.cell(row=row, column=6, value=stock_in.total_price or 0)
                ws.cell(row=row, column=7, value=stock_in.invoice_number or "")
                ws.cell(row=row, column=8, value=gregorian_to_persian_str(stock_in.manual_date, "%Y/%m/%d") if stock_in.manual_date else "")
                ws.cell(row=row, column=9, value=stock_in.notes or "")
            
            # ذخیره فایل
            filename = f"ورودی_انبار_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join("media", "excel_reports", filename)
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            wb.save(filepath)
            
            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
                
        except Exception as e:
            messages.error(request, f'خطا در صدور فایل: {str(e)}')
            return redirect('admin:inventory_stockin_changelist')
    
    export_stock_in_excel.short_description = "صدور ورودی‌های انتخاب شده به Excel"

@admin.register(StockOut)
class StockOutAdmin(admin.ModelAdmin):
    list_display = ['warehouse', 'material_type', 'customer', 'supplier', 'quantity', 'unit_price', 'total_price', 'persian_manual_date', 'persian_created_at']
    list_filter = ['warehouse', 'material_type', 'customer', 'supplier', 'created_at', 'manual_date']
    search_fields = ['warehouse__name', 'material_type__name', 'customer__name', 'supplier__name', 'invoice_number']
    readonly_fields = ['total_price', 'created_at']
    date_hierarchy = 'created_at'
    actions = ['export_stock_out_excel']
    
    def persian_manual_date(self, obj):
        if obj.manual_date:
            return gregorian_to_persian_str(obj.manual_date, "%Y/%m/%d")
        return "-"
    persian_manual_date.short_description = 'تاریخ خروج دستی (شمسی)'
    
    def persian_created_at(self, obj):
        return gregorian_to_persian_datetime_str(obj.created_at, "%Y/%m/%d %H:%M")
    persian_created_at.short_description = 'تاریخ ثبت (شمسی)'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.admin_site.admin_view(self.upload_excel_view), name='inventory_stockout_upload_excel'),
            path('download-template/', self.admin_site.admin_view(self.download_template_view), name='inventory_stockout_download_template'),
        ]
        return custom_urls + urls
    
    def upload_excel_view(self, request):
        if request.method == 'POST':
            uploaded_file = request.FILES.get('excel_file')
            if uploaded_file:
                try:
                    # ذخیره فایل موقت
                    temp_path = f'/tmp/stock_out_admin_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                    with open(temp_path, 'wb+') as destination:
                        for chunk in uploaded_file.chunks():
                            destination.write(chunk)
                    
                    # پردازش فایل
                    results = import_stock_out_excel(temp_path, request.user)
                    
                    # حذف فایل موقت
                    os.remove(temp_path)
                    
                    # نمایش نتایج
                    success_count = len(results.get('success', []))
                    error_count = len(results.get('errors', []))
                    
                    if success_count > 0:
                        messages.success(request, f'{success_count} رکورد با موفقیت وارد شد.')
                    
                    if error_count > 0:
                        for error in results['errors'][:5]:  # نمایش 5 خطای اول
                            messages.error(request, error)
                        if error_count > 5:
                            messages.warning(request, f'و {error_count - 5} خطای دیگر وجود دارد.')
                    
                    return redirect('admin:inventory_stockout_changelist')
                    
                except Exception as e:
                    messages.error(request, f'خطا در پردازش فایل: {str(e)}')
            else:
                messages.error(request, 'فایل انتخاب نشده است.')
        
        return render(request, 'admin/inventory/stockout/upload_excel.html', {
            'title': 'آپلود Excel خروجی انبار',
            'opts': self.model._meta,
        })
    
    def download_template_view(self, request):
        try:
            file_path = create_stock_out_template()
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{os.path.basename(file_path)}"'
                return response
        except Exception as e:
            messages.error(request, f'خطا در ایجاد قالب: {str(e)}')
            return redirect('admin:inventory_stockout_changelist')
    
    def export_stock_out_excel(self, request, queryset):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "خروجی انبار"
            
            # هدرها
            headers = [
                "انبار", "نام کالا", "نام مشتری", "هویت کالای خروجی", "مقدار", "قیمت واحد", 
                "قیمت کل", "شماره بارنامه", "تاریخ خروج دستی", "یادداشت‌ها"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="C5504B", end_color="C5504B", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # داده‌ها
            for row, stock_out in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=stock_out.warehouse.name)
                ws.cell(row=row, column=2, value=stock_out.material_type.name)
                ws.cell(row=row, column=3, value=stock_out.customer.name)
                ws.cell(row=row, column=4, value=stock_out.quantity or 0)
                ws.cell(row=row, column=5, value=stock_out.unit_price or 0)
                ws.cell(row=row, column=6, value=stock_out.total_price or 0)
                ws.cell(row=row, column=7, value=stock_out.invoice_number or "")
                ws.cell(row=row, column=8, value=gregorian_to_persian_str(stock_out.manual_date, "%Y/%m/%d") if stock_out.manual_date else "")
                ws.cell(row=row, column=9, value=stock_out.notes or "")
            
            # ذخیره فایل
            filename = f"خروجی_انبار_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join("media", "excel_reports", filename)
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            wb.save(filepath)
            
            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
                
        except Exception as e:
            messages.error(request, f'خطا در صدور فایل: {str(e)}')
            return redirect('admin:inventory_stockout_changelist')
    
    export_stock_out_excel.short_description = "صدور خروجی‌های انتخاب شده به Excel"

@admin.register(StockTransfer)
class StockTransferAdmin(admin.ModelAdmin):
    list_display = ['source_warehouse', 'destination_warehouse', 'material_type', 'quantity', 'created_by', 'persian_created_at']
    list_filter = ['source_warehouse', 'destination_warehouse', 'created_at', 'material_type']
    search_fields = ['source_warehouse__name', 'destination_warehouse__name', 'material_type__name', 'notes']
    readonly_fields = ['created_by', 'created_at']
    date_hierarchy = 'created_at'
    actions = ['export_stock_transfer_excel']
    
    def persian_created_at(self, obj):
        return gregorian_to_persian_datetime_str(obj.created_at, "%Y/%m/%d %H:%M")
    persian_created_at.short_description = 'تاریخ انتقال (شمسی)'
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.upload_excel_view, name='upload_excel'),
            path('download-template/', self.download_template_view, name='download_template'),
        ]
        return custom_urls + urls
    
    def upload_excel_view(self, request):
        if request.method == 'POST':
            if 'excel_file' not in request.FILES:
                messages.error(request, "فایل انتخاب نشده است")
                return redirect('admin:inventory_stocktransfer_changelist')
            
            excel_file = request.FILES['excel_file']
            
            if not excel_file.name.endswith('.xlsx'):
                messages.error(request, "فقط فایل‌های Excel (.xlsx) قابل قبول هستند")
                return redirect('admin:inventory_stocktransfer_changelist')
            
            try:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                    for chunk in excel_file.chunks():
                        temp_file.write(chunk)
                    temp_path = temp_file.name
                
                results = import_stock_transfer_excel(temp_path, request.user)
                os.unlink(temp_path)
                
                if results['success']:
                    for success_msg in results['success']:
                        messages.success(request, success_msg)
                
                if results['errors']:
                    for error_msg in results['errors']:
                        messages.error(request, error_msg)
                
            except Exception as e:
                messages.error(request, f"خطا در پردازش فایل: {str(e)}")
            
            return redirect('admin:inventory_stocktransfer_changelist')
        
        return render(request, 'admin/inventory/stocktransfer/upload_excel.html')
    
    def download_template_view(self, request):
        try:
            file_path = create_stock_transfer_template()
            with open(file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="قالب_انتقال_انبار.xlsx"'
                return response
        except Exception as e:
            messages.error(request, f"خطا در ایجاد قالب: {str(e)}")
            return redirect('admin:inventory_stocktransfer_changelist')
    
    def export_stock_transfer_excel(self, request, queryset):
        """صدور انتقالات انبار به Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "انتقالات انبار"
            
            # تعریف هدرها
            headers = [
                "انبار مبدا",
                "انبار مقصد",
                "نام کالا",
                "مقدار",
                "یادداشت‌ها",
                "ثبت کننده",
                "تاریخ انتقال"
            ]
            
            # اضافه کردن هدرها
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # تنظیم عرض ستون‌ها
            column_widths = [20, 20, 20, 15, 30, 20, 20]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            # اضافه کردن داده‌ها
            for row, transfer in enumerate(queryset, 2):
                ws.cell(row=row, column=1, value=transfer.source_warehouse.name)
                ws.cell(row=row, column=2, value=transfer.destination_warehouse.name)
                ws.cell(row=row, column=3, value=transfer.material_type.name)
                ws.cell(row=row, column=4, value=transfer.quantity or 0)
                ws.cell(row=row, column=5, value=transfer.notes)
                ws.cell(row=row, column=6, value=transfer.created_by.username)
                ws.cell(row=row, column=7, value=gregorian_to_persian_datetime_str(transfer.created_at, "%Y/%m/%d %H:%M"))
            
            # ذخیره فایل
            filename = f"انتقالات_انبار_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join("media", "excel_reports", filename)
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            wb.save(filepath)
            
            with open(filepath, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
                
        except Exception as e:
            messages.error(request, f"خطا در صدور فایل: {str(e)}")
            return redirect('admin:inventory_stocktransfer_changelist')
    
    export_stock_transfer_excel.short_description = "صدور انتقالات انتخاب شده به Excel"
