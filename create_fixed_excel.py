#!/usr/bin/env python3
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Create corrected data with Persian dates
corrected_data = [
    {
        "انبار": "انبار اصلی",
        "نوع عملیات": "ورودی",
        "نام کالا": "میلگرد 16",
        "هویت کالا/نام مشتری": "شرکت آهن آلات تهران",
        "مقدار": 1000,
        "قیمت واحد": 15000,
        "شماره بارنامه": "BR001",
        "تاریخ (YYYY-MM-DD)": "1404-01-26",  # Persian date
        "یادداشت‌ها": "ورودی اولیه میلگرد"
    },
    {
        "انبار": "انبار اصلی",
        "نوع عملیات": "خروجی",
        "نام کالا": "میلگرد 16",
        "هویت کالا/نام مشتری": "شرکت ساختمانی آسمان",
        "مقدار": 200,
        "قیمت واحد": 18000,
        "شماره بارنامه": "BR002",
        "تاریخ (YYYY-MM-DD)": "1404-01-27",  # Persian date
        "یادداشت‌ها": "خروجی اولیه میلگرد"
    },
    {
        "انبار": "انبار اصلی",
        "نوع عملیات": "ورودی",
        "نام کالا": "ورق فولادی",
        "هویت کالا/نام مشتری": "کارخانه فولاد اصفهان",
        "مقدار": 500,
        "قیمت واحد": 25000,
        "شماره بارنامه": "BR003",
        "تاریخ (YYYY-MM-DD)": "1404-01-28",  # Persian date
        "یادداشت‌ها": "ورودی ورق فولادی"
    },
    {
        "انبار": "انبار اصلی",
        "نوع عملیات": "خروجی",
        "نام کالا": "ورق فولادی",
        "هویت کالا/نام مشتری": "پروژه برج تهران",
        "مقدار": 100,
        "قیمت واحد": 28000,
        "شماره بارنامه": "BR004",
        "تاریخ (YYYY-MM-DD)": "1404-01-29",  # Persian date
        "یادداشت‌ها": "خروجی ورق فولادی"
    },
    {
        "انبار": "انبار فرعی",
        "نوع عملیات": "ورودی",
        "نام کالا": "نبشی 50",
        "هویت کالا/نام مشتری": "کارخانه فولاد کرج",
        "مقدار": 300,
        "قیمت واحد": 12000,
        "شماره بارنامه": "BR005",
        "تاریخ (YYYY-MM-DD)": "1404-01-30",  # Persian date
        "یادداشت‌ها": "ورودی نبشی"
    }
]

# Create DataFrame
df = pd.DataFrame(corrected_data)

# Create Excel file with proper formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ورودی و خروجی انبار"

# Add headers with styling
headers = list(corrected_data[0].keys())
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add data with styling
for row_idx, row_data in enumerate(corrected_data, 2):
    for col_idx, (key, value) in enumerate(row_data.items(), 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Color coding for operation types
        if col_idx == 2:  # Operation type column
            if value == "ورودی":
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                cell.font = Font(bold=True, color="006400")
            elif value == "خروجی":
                cell.fill = PatternFill(start_color="F8E8E8", end_color="F8E8E8", fill_type="solid")
                cell.font = Font(bold=True, color="8B0000")
        
        # Highlight Persian dates
        if col_idx == 8:  # Date column
            if str(value).startswith("1404"):
                cell.fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
                cell.font = Font(bold=True, color="FF8C00")

# Set column widths
column_widths = [15, 15, 20, 25, 15, 15, 20, 20, 30]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# Add a summary sheet
summary_sheet = wb.create_sheet("خلاصه")
summary_sheet['A1'] = "خلاصه عملیات انبار"
summary_sheet['A1'].font = Font(bold=True, size=16)

summary_sheet['A3'] = "تعداد کل ردیف‌ها:"
summary_sheet['B3'] = len(corrected_data)
summary_sheet['A4'] = "تعداد ورودی‌ها:"
summary_sheet['B4'] = len([d for d in corrected_data if d['نوع عملیات'] == 'ورودی'])
summary_sheet['A5'] = "تعداد خروجی‌ها:"
summary_sheet['B5'] = len([d for d in corrected_data if d['نوع عملیات'] == 'خروجی'])

# Save the file
filename = "قالب_یکپارچه_صحیح_با_تاریخ_فارسی.xlsx"
wb.save(filename)
print(f"✅ فایل Excel اصلاح شده ایجاد شد: {filename}")
print(f"📊 تعداد ردیف‌ها: {len(corrected_data)}")
print(f"📥 ورودی‌ها: {len([d for d in corrected_data if d['نوع عملیات'] == 'ورودی'])}")
print(f"📤 خروجی‌ها: {len([d for d in corrected_data if d['نوع عملیات'] == 'خروجی'])}")
print(f"📅 تاریخ‌های فارسی: {len([d for d in corrected_data if str(d['تاریخ (YYYY-MM-DD)']).startswith('1404')])}")
print(f"📁 مسیر فایل: {os.path.abspath(filename)}")
