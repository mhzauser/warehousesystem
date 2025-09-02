#!/usr/bin/env python3
import pandas as pd
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Create test data with Persian dates
test_data = [
    {
        "انبار": "انبار اصلی",
        "نوع عملیات": "ورودی",
        "نام کالا": "میلگرد 16",
        "هویت کالا/نام مشتری": "شرکت آهن آلات تهران",
        "مقدار": 1000,
        "قیمت واحد": 15000,
        "شماره بارنامه": "BR001",
        "تاریخ (YYYY-MM-DD)": "1404-01-26",  # Persian date
        "یادداشت‌ها": "ورودی اولیه"
    },
    {
        "انبار": "انبار اصلی",
        "نوع عملیات": "خروجی",
        "نام کالا": "میلگرد 16",
        "هویت کالا/نام مشتری": "شرکت ساختمانی آسمان",
        "مقدار": 200,
        "قیمت واحد": 18000,
        "شماره بارنامه": "BR002",
        "تاریخ (YYYY-MM-DD)": "1404-01-27",  # Persian date
        "یادداشت‌ها": "خروجی اولیه"
    },
    {
        "انبار": "انبار اصلی",
        "نوع عملیات": "ورودی",
        "نام کالا": "ورق فولادی",
        "هویت کالا/نام مشتری": "کارخانه فولاد اصفهان",
        "مقدار": 500,
        "قیمت واحد": 25000,
        "شماره بارنامه": "BR003",
        "تاریخ (YYYY-MM-DD)": "1404-01-28",  # Persian date
        "یادداشت‌ها": "ورودی دوم"
    },
    {
        "انبار": "انبار اصلی",
        "نوع عملیات": "خروجی",
        "نام کالا": "ورق فولادی",
        "هویت کالا/نام مشتری": "پروژه برج تهران",
        "مقدار": 100,
        "قیمت واحد": 28000,
        "شماره بارنامه": "BR004",
        "تاریخ (YYYY-MM-DD)": "1404-01-29",  # Persian date
        "یادداشت‌ها": "خروجی دوم"
    }
]

# Create DataFrame
df = pd.DataFrame(test_data)

# Create Excel file with proper formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ورودی و خروجی انبار"

# Add headers with styling
headers = list(test_data[0].keys())
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add data
for row_idx, row_data in enumerate(test_data, 2):
    for col_idx, (key, value) in enumerate(row_data.items(), 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Color coding for operation types
        if col_idx == 2:  # Operation type column
            if value == "ورودی":
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            elif value == "خروجی":
                cell.fill = PatternFill(start_color="F8E8E8", end_color="F8E8E8", fill_type="solid")

# Set column widths
column_widths = [15, 15, 20, 25, 15, 15, 20, 20, 30]
for col, width in enumerate(column_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

# Save the file
filename = "test_persian_dates.xlsx"
wb.save(filename)
print(f"Test Excel file created: {filename}")
print(f"File contains {len(test_data)} rows with Persian dates")
print("Columns:", headers)
